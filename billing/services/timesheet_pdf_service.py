"""
稼働報告PDF生成サービス
ExcelファイルをGoogle Drive経由でPDFに変換してアップロードする。

フォルダ構成:
  パートナー管理/{パートナー名}/稼働報告書/  ← WorkReport (パートナー稼働報告)
  クライアント管理/{クライアント名}/稼働報告書/  ← MonthlyTimesheet (自社勤怠)
"""
import logging

logger = logging.getLogger(__name__)


def _detect_target_sheet_name(excel_data):
    """
    Excel解析と同じロジックで対象シート名を検出する。
    IGNORE_SHEET_KEYWORDS（サンプル、祝日等）を除外し、
    日付データが最も多いシートを返す。
    """
    import io
    import openpyxl
    from datetime import datetime

    IGNORE_KEYWORDS = ['サンプル', 'テスト', 'sample', 'test', '祝日', 'holiday', 'template', 'テンプレ']

    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_data), data_only=True, keep_vba=False)
        candidates = []
        for sheet_name in wb.sheetnames:
            if any(kw in sheet_name.lower() for kw in IGNORE_KEYWORDS):
                continue
            ws = wb[sheet_name]
            date_count = sum(
                1 for row_idx in range(1, min(ws.max_row + 1, 45))
                if isinstance(ws.cell(row=row_idx, column=2).value, datetime)
            )
            if date_count >= 10:
                candidates.append((date_count, sheet_name))
        
        if candidates:
            candidates.sort(key=lambda x: x[0], reverse=True)
            return candidates[0][1]
        
        # フォールバック: 2番目のシート（よくあるパターン: サンプル/本体/祝日）
        if len(wb.sheetnames) > 1:
            return wb.sheetnames[1]
        return wb.sheetnames[0]
    except Exception as e:
        logger.warning(f'[TimesheetPDF] シート検出エラー: {e}')
        return None


def _convert_excel_to_pdf_via_drive(excel_data, filename, target_sheet_name=None):
    """
    Google Drive APIを使ってExcelをPDFに変換する。
    1. 対象シートだけ残したExcelを作成
    2. Google Sheetsとしてアップロード（変換）
    3. files().export() でPDFエクスポート
    4. 一時ファイルを削除
    
    Args:
        excel_data: Excelバイトデータ
        filename: ファイル名
        target_sheet_name: PDF化する対象シート名（Noneなら全シート）
    
    Returns:
        bytes: PDFバイトデータ, or None
    """
    from core.services.google_drive_service import _get_drive_service
    from googleapiclient.http import MediaInMemoryUpload
    import io
    import openpyxl

    service = _get_drive_service()
    if not service:
        return None

    try:
        # 対象シートだけ残したExcelを作成
        upload_data = excel_data
        if target_sheet_name:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(excel_data), keep_vba=False)
                sheets_to_remove = [s for s in wb.sheetnames if s != target_sheet_name]
                for s in sheets_to_remove:
                    del wb[s]
                buf = io.BytesIO()
                wb.save(buf)
                upload_data = buf.getvalue()
                logger.info(f'[TimesheetPDF] 対象シート「{target_sheet_name}」のみのExcel作成 ({len(upload_data)} bytes)')
            except Exception as e:
                logger.warning(f'[TimesheetPDF] シート絞り込み失敗、全シートで続行: {e}')

        # 共有ドライブ内にアップロード
        from django.conf import settings as django_settings
        root_folder_id = getattr(django_settings, 'GOOGLE_DRIVE_ROOT_FOLDER_ID', '')

        media = MediaInMemoryUpload(
            upload_data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        file_metadata = {
            'name': filename,
            'mimeType': 'application/vnd.google-apps.spreadsheet',
        }
        if root_folder_id:
            file_metadata['parents'] = [root_folder_id]
        uploaded = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id',
            supportsAllDrives=True,
        ).execute()
        temp_file_id = uploaded.get('id')

        # PDFエクスポート（Drive API）
        pdf_data = service.files().export(
            fileId=temp_file_id,
            mimeType='application/pdf',
        ).execute()

        # 一時ファイルを削除
        try:
            service.files().delete(fileId=temp_file_id, supportsAllDrives=True).execute()
        except Exception as e:
            logger.warning(f'[TimesheetPDF] 一時ファイル削除失敗: {e}')

        logger.info(f'[TimesheetPDF] Excel→PDF変換成功: {filename} ({len(pdf_data)} bytes)')
        return pdf_data

    except Exception as e:
        logger.error(f'[TimesheetPDF] Excel→PDF変換エラー: {e}')
        return None


def generate_timesheet_pdf(timesheet):
    """
    MonthlyTimesheetのExcelをPDFに変換してGoogle Driveにアップロードする。
    → クライアント管理/{クライアント名}/稼働報告書/
    """
    from core.services.google_drive_service import upload_document

    # Excelファイルの取得
    if not timesheet.excel_file:
        logger.warning(f'[TimesheetPDF] Excelファイルなし: {timesheet.worker_name}')
        return timesheet

    timesheet.excel_file.seek(0)
    excel_data = timesheet.excel_file.read()

    month_str = timesheet.target_month.strftime('%Y%m')
    clean_name = timesheet.worker_name.replace('\u3000', '').replace(' ', '')
    pdf_filename = f'稼働報告_{clean_name}_{month_str}.pdf'

    # Excel → PDF変換（対象シートを自動検出）
    target_sheet = _detect_target_sheet_name(excel_data)
    logger.info(f'[TimesheetPDF] 対象シート: {target_sheet}')
    pdf_bytes = _convert_excel_to_pdf_via_drive(excel_data, f'temp_{pdf_filename}', target_sheet)
    if not pdf_bytes:
        logger.warning(f'[TimesheetPDF] PDF変換失敗: {timesheet.worker_name}')
        return timesheet

    # ローカルにPDF保存
    from django.core.files.base import ContentFile
    timesheet.pdf_file.save(
        f'timesheet_{timesheet.pk}_{month_str}.pdf',
        ContentFile(pdf_bytes),
        save=True,
    )
    logger.info(f'[TimesheetPDF] ローカルPDF保存: {timesheet.pdf_file.name}')

    # Google Driveにアップロード（クライアント管理）
    customer_name = timesheet.order.customer.name if timesheet.order else '未分類'
    file_id, web_link = upload_document(
        management_type='client',
        company_name=customer_name,
        doc_type='work_report',
        filename=pdf_filename,
        pdf_bytes=pdf_bytes,
    )
    if file_id:
        timesheet.drive_file_id = file_id
        timesheet.save(update_fields=['drive_file_id'])
        logger.info(f'[TimesheetPDF] Drive保存完了: {pdf_filename} (ID: {file_id})')

    return timesheet


def generate_workreport_pdf(work_report):
    """
    WorkReportのExcelをPDFに変換してGoogle Driveにアップロードする。
    → パートナー管理/{パートナー名}/稼働報告書/
    """
    from core.services.google_drive_service import upload_document

    # Excelファイルの取得
    if not work_report.file:
        logger.warning(f'[WorkReportPDF] Excelファイルなし: {work_report.worker_name}')
        return work_report

    work_report.file.seek(0)
    excel_data = work_report.file.read()

    month_str = work_report.target_month.strftime('%Y%m') if work_report.target_month else 'unknown'
    pdf_filename = f'稼働報告_{work_report.worker_name}_{month_str}.pdf'

    # Excel → PDF変換（対象シートを自動検出）
    target_sheet = _detect_target_sheet_name(excel_data)
    logger.info(f'[WorkReportPDF] 対象シート: {target_sheet}')
    pdf_bytes = _convert_excel_to_pdf_via_drive(excel_data, f'temp_{pdf_filename}', target_sheet)
    if not pdf_bytes:
        logger.warning(f'[WorkReportPDF] PDF変換失敗: {work_report.worker_name}')
        return work_report

    # ローカルにPDF保存
    from django.core.files.base import ContentFile
    work_report.pdf_file.save(
        f'workreport_{work_report.pk}_{month_str}.pdf',
        ContentFile(pdf_bytes),
        save=True,
    )
    logger.info(f'[WorkReportPDF] ローカルPDF保存: {work_report.pdf_file.name}')

    # Google Driveにアップロード（パートナー管理）
    partner_name = '未分類'
    if work_report.order and work_report.order.partner:
        partner_name = work_report.order.partner.name

    file_id, web_link = upload_document(
        management_type='partner',
        company_name=partner_name,
        doc_type='work_report',
        filename=pdf_filename,
        pdf_bytes=pdf_bytes,
    )
    if file_id:
        work_report.drive_file_id = file_id
        work_report.save(update_fields=['drive_file_id'])
        logger.info(f'[WorkReportPDF] Drive保存完了: {pdf_filename} (ID: {file_id})')

    return work_report
