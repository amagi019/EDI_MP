"""
稼働報告書 Excel 自動読取りサービス

Excelブック（.xlsx / .xlsm）を受け取り、以下を自動判別して読み取る:
  - 稼働報告シートの特定（サンプル/祝日シートは除外）
  - B列の日付, D列の作業時間(分), T+U列の作業時間(h:mm)
  - 合計時間(E41), 稼働日数(E42)
  - オートシェイプ / ファイル名からの作業者名取得
  - 土日祝の稼働チェック
"""
import io
import re
import logging
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, date, timedelta
from collections import Counter
from decimal import Decimal

import openpyxl
import jpholiday

logger = logging.getLogger(__name__)

# シート名に含まれていたら除外するキーワード
IGNORE_SHEET_KEYWORDS = ['サンプル', 'テスト', 'sample', 'test', '祝日', 'holiday', 'template', 'テンプレ']


# ──────────────────────────────────────────
# メインAPI
# ──────────────────────────────────────────

def auto_detect_and_parse(file_obj, original_filename=''):
    """
    Excelファイルを自動判別してパースする。

    Args:
        file_obj: ファイルオブジェクト (seekable)
        original_filename: 元のファイル名

    Returns:
        dict: {
            'worker_name': str,
            'target_month': date or None,
            'total_hours': Decimal,
            'work_days': int,
            'daily_data': [{'date': 'YYYY-MM-DD', 'hours': float, 'start': '9:00', 'end': '18:00'}, ...],
            'alerts': [{'date': 'YYYY-MM-DD', 'type': 'weekend'|'holiday', 'hours': float, 'day_name': str, 'holiday_name': str}, ...],
            'error': str or None,
        }
    """
    try:
        file_obj.seek(0)
        file_bytes = file_obj.read()
        file_obj.seek(0)

        # data_only=True で計算結果を取得
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), data_only=True, keep_vba=False
        )

        # 1. 稼働報告シートの特定
        ws = _detect_work_sheet(wb)
        if ws is None:
            return _error_result('稼働報告データのあるシートが見つかりませんでした。')

        # 2. 日付列と日付データ範囲の検出
        date_col, data_rows = _detect_date_column(ws)
        if date_col is None:
            return _error_result('日付データの列を検出できませんでした。')

        # 3. 作業時間列の検出
        hours_col, hours_unit = _detect_hours_column(ws, date_col, data_rows)
        if hours_col is None:
            return _error_result('作業時間データの列を検出できませんでした。')

        # 4. 日別データの抽出
        daily_data = _extract_daily_data(ws, date_col, hours_col, hours_unit, data_rows)

        # 5. 作業月の検出（B列データが最優先）
        target_month = _detect_target_month(daily_data, ws, original_filename)

        # 6. 合計時間・稼働日数の検出
        total_hours, work_days = _detect_totals(ws, daily_data, data_rows)

        # 7. 作業者名の取得
        worker_name, name_mismatch = _extract_worker_name(file_bytes, ws, original_filename)

        # 8. 土日祝・稼働時間チェック
        holiday_dates = _parse_holiday_sheet(wb)
        alerts = check_work_alerts(daily_data, holiday_dates)

        return {
            'worker_name': worker_name,
            'target_month': target_month,
            'total_hours': total_hours,
            'work_days': work_days,
            'daily_data': daily_data,
            'alerts': alerts,
            'name_mismatch_warning': name_mismatch,
            'sheet_name': ws.title,
            'error': None,
        }

    except Exception as e:
        logger.exception(f'Excel解析エラー: {e}')
        return _error_result(f'ファイルの解析中にエラーが発生しました: {e}')


def _error_result(message):
    return {
        'worker_name': '',
        'target_month': None,
        'total_hours': Decimal('0'),
        'work_days': 0,
        'daily_data': [],
        'alerts': [],
        'error': message,
    }


# ──────────────────────────────────────────
# Step 1: シート検出
# ──────────────────────────────────────────

def _detect_work_sheet(wb):
    """稼働報告データのあるシートを特定する。「サンプル」「祝日」等は除外。"""
    candidates = []

    for sheet_name in wb.sheetnames:
        # 除外キーワードチェック
        if any(kw in sheet_name.lower() for kw in IGNORE_SHEET_KEYWORDS):
            continue

        ws = wb[sheet_name]
        score = _score_sheet(ws)
        if score > 0:
            candidates.append((score, sheet_name))

    if not candidates:
        # 除外対象も含めて再チェック（全シートがサンプルの場合のフォールバック）
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            score = _score_sheet(ws)
            if score > 0:
                candidates.append((score, sheet_name))

    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        best_name = candidates[0][1]
        logger.info(f'[Excel] 検出シート: {best_name} (スコア: {candidates[0][0]})')
        return wb[best_name]

    return None


def _score_sheet(ws):
    """シートが稼働報告データを含んでいるかスコアリングする。"""
    score = 0
    date_count = 0

    # B列（2列目）の最初の40行をチェック
    for row_idx in range(1, min(ws.max_row + 1, 45)):
        cell = ws.cell(row=row_idx, column=2)
        if isinstance(cell.value, datetime):
            date_count += 1

    # 20日以上の日付データがあれば高スコア
    if date_count >= 20:
        score += date_count
    elif date_count >= 10:
        score += date_count // 2

    return score


# ──────────────────────────────────────────
# Step 2: 日付列検出
# ──────────────────────────────────────────

def _detect_date_column(ws):
    """日付データが連続して存在する列とデータ行範囲を検出する。"""
    best_col = None
    best_rows = []
    best_count = 0

    # A〜E列をチェック（通常B列）
    for col_idx in range(1, 6):
        rows = []
        for row_idx in range(1, min(ws.max_row + 1, 45)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, datetime):
                rows.append(row_idx)

        if len(rows) > best_count:
            best_count = len(rows)
            best_col = col_idx
            best_rows = rows

    if best_count >= 20:
        return best_col, best_rows

    return None, []


# ──────────────────────────────────────────
# Step 3: 作業時間列検出
# ──────────────────────────────────────────

def _detect_hours_column(ws, date_col, data_rows):
    """
    作業時間の列を検出する。

    パターン1: D列のような「分単位」の列（数式 = (終了-開始-休憩)*60）
    パターン2: T列のような「時間」の列
    """
    if not data_rows:
        return None, None

    # ヘッダー行を検出（data_rowsの最初の行の1つ前）
    header_row = data_rows[0] - 1 if data_rows[0] > 1 else 1

    # 候補列を探す（日付列の右側）
    candidates = []

    for col_idx in range(date_col + 1, min(ws.max_column + 1, 25)):
        values = []
        for row_idx in data_rows:
            v = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(v, (int, float)):
                values.append(v)

        if not values:
            continue

        non_zero = [v for v in values if v > 0]
        if not non_zero:
            continue

        max_val = max(non_zero)
        avg_val = sum(non_zero) / len(non_zero)

        # ヘッダーテキスト
        header = str(ws.cell(row=header_row, column=col_idx).value or '').lower()

        # パターン1: 分単位（60〜1440の範囲が多い）
        if max_val <= 1440 and avg_val >= 60:
            score = len(non_zero)
            if '合計' in header or 'min' in header:
                score += 20
            candidates.append((score, col_idx, 'minutes'))

        # パターン2: 時間単位（0〜24の範囲）
        elif max_val <= 24 and avg_val >= 1:
            score = len(non_zero)
            if '合計' in header or '時間' in header or 'h' in header:
                score += 20
            candidates.append((score, col_idx, 'hours'))

    if candidates:
        # 分単位の列を優先（D列の数式で正確に計算されているため）
        minute_candidates = [c for c in candidates if c[2] == 'minutes']
        if minute_candidates:
            minute_candidates.sort(key=lambda x: x[0], reverse=True)
            return minute_candidates[0][1], 'minutes'

        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1], candidates[0][2]

    return None, None


# ──────────────────────────────────────────
# Step 4: 日別データ抽出
# ──────────────────────────────────────────

def _extract_daily_data(ws, date_col, hours_col, hours_unit, data_rows):
    """日別の稼働データを抽出する。"""
    daily_data = []

    for row_idx in data_rows:
        date_val = ws.cell(row=row_idx, column=date_col).value
        if not isinstance(date_val, datetime):
            continue

        d = date_val.date() if isinstance(date_val, datetime) else date_val

        # 作業時間
        hours_val = ws.cell(row=row_idx, column=hours_col).value or 0
        if hours_unit == 'minutes':
            hours = round(float(hours_val) / 60, 2)
        else:
            hours = round(float(hours_val), 2)

        # 開始・終了時刻の検出（L,N列 → P,R列パターン）
        start_str = ''
        end_str = ''
        # L列(12), N列(14) → 開始
        l_val = ws.cell(row=row_idx, column=12).value
        n_val = ws.cell(row=row_idx, column=14).value
        if l_val is not None and hours > 0:
            start_str = f"{int(l_val)}:{int(n_val or 0):02d}"
        # P列(16), R列(18) → 終了
        p_val = ws.cell(row=row_idx, column=16).value
        r_val = ws.cell(row=row_idx, column=18).value
        if p_val is not None and hours > 0:
            end_str = f"{int(p_val)}:{int(r_val or 0):02d}"

        day_names = ['月', '火', '水', '木', '金', '土', '日']
        daily_data.append({
            'date': d.isoformat(),
            'day_name': day_names[d.weekday()],
            'hours': hours,
            'start': start_str,
            'end': end_str,
        })

    return daily_data


# ──────────────────────────────────────────
# Step 5: 作業月検出
# ──────────────────────────────────────────

def _detect_target_month(daily_data, ws, filename):
    """
    作業月を検出する。優先順位:
      1. B列の日付データ（最も信頼性が高い）
      2. D3セル
      3. ファイル名
    """
    # 1. 日別データから月を推定
    if daily_data:
        months = Counter()
        for d in daily_data:
            dt = date.fromisoformat(d['date'])
            months[(dt.year, dt.month)] += 1
        if months:
            most_common = months.most_common(1)[0][0]
            return date(most_common[0], most_common[1], 1)

    # 2. D3セル
    d3 = ws.cell(row=3, column=4).value
    if isinstance(d3, datetime):
        return d3.date().replace(day=1)

    # 3. ファイル名から
    m = re.search(r'(\d{4})[年/\-](\d{1,2})', filename)
    if m:
        return date(int(m.group(1)), int(m.group(2)), 1)

    return None


# ──────────────────────────────────────────
# Step 6: 合計・稼働日数
# ──────────────────────────────────────────

def _detect_totals(ws, daily_data, data_rows):
    """合計時間と稼働日数を検出する。"""
    # 日別データから計算
    total_from_data = sum(d['hours'] for d in daily_data)
    days_from_data = sum(1 for d in daily_data if d['hours'] > 0)

    # E41（時間合計）、E42（稼働日数）もチェック
    e41 = ws.cell(row=41, column=5).value
    e42 = ws.cell(row=42, column=5).value

    # E41がある場合はそちらを信用（数式の計算結果）
    if e41 and isinstance(e41, (int, float)) and e41 > 0:
        total_hours = Decimal(str(e41))
    else:
        total_hours = Decimal(str(round(total_from_data, 2)))

    if e42 and isinstance(e42, (int, float)) and e42 > 0:
        work_days = int(e42)
    else:
        work_days = days_from_data

    return total_hours, work_days


# ──────────────────────────────────────────
# Step 7: 作業者名取得
# ──────────────────────────────────────────

def _extract_worker_name(file_bytes, ws, filename):
    """
    作業者名を取得する。優先順位:
      1. 検出シートに紐づくDrawingMLの「氏名：」パターン
      2. ファイル名の（名前）パターン
    ファイル名とDrawingの氏名が不一致の場合は警告を返す。

    Returns:
        tuple: (worker_name, mismatch_warning or '')
    """
    drawing_name = _extract_name_from_drawings(file_bytes, ws.title)

    # ファイル名から抽出
    filename_name = ''
    m = re.search(r'[（(](.+?)[）)]', filename)
    if m:
        name_candidate = m.group(1)
        if not any(kw in name_candidate for kw in ['報告', '作業', 'コピー', 'copy']):
            filename_name = name_candidate

    # 両方取得できた場合にマッチング
    mismatch = ''
    if drawing_name and filename_name:
        # 空白除去して比較
        dn = drawing_name.replace('\u3000', '').replace(' ', '')
        fn = filename_name.replace('\u3000', '').replace(' ', '')
        if dn != fn:
            mismatch = f'ファイル名の氏名「{filename_name}」とシート内の氏名「{drawing_name}」が一致しません'

    # Drawing名を優先、なければファイル名
    worker_name = drawing_name or filename_name
    return worker_name, mismatch


def _extract_name_from_drawings(file_bytes, target_sheet_name=None):
    """
    xlsxのZIP内のDrawingMLから作業者名を抽出する。
    target_sheet_name指定時は、そのシートに紐づくDrawingのみ検索。
    """
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes), 'r') as z:
            # シート名→sheetN.xml→drawingN.xml の対応を構築
            target_drawings = None
            if target_sheet_name:
                target_drawings = _get_drawings_for_sheet(z, target_sheet_name)

            if target_drawings:
                drawing_files = target_drawings
            else:
                # フォールバック: 全drawingファイルを検索
                drawing_files = [
                    f for f in z.namelist()
                    if f.startswith('xl/drawings/drawing') and f.endswith('.xml')
                ]

            for df in drawing_files:
                content = z.read(df)
                try:
                    root = ET.fromstring(content)
                    ns_a = '{http://schemas.openxmlformats.org/drawingml/2006/main}'
                    texts = []
                    for t_elem in root.iter(f'{ns_a}t'):
                        if t_elem.text:
                            texts.append(t_elem.text.strip())

                    for text in texts:
                        # 「氏名：」「氏　名：」パターン
                        m = re.search(r'氏\s*名\s*[：:][\s　]*(.+)', text)
                        if m:
                            name = m.group(1).strip()
                            if name:
                                return name
                except ET.ParseError:
                    continue
    except (zipfile.BadZipFile, Exception) as e:
        logger.debug(f'DrawingML解析スキップ: {e}')

    return ''


def _get_drawings_for_sheet(z, sheet_name):
    """シート名に対応するdrawingファイルパスを返す。"""
    try:
        # workbook.xmlからシート名→rIdの対応を取得
        wb_content = z.read('xl/workbook.xml')
        wb_root = ET.fromstring(wb_content)
        ns_main = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
        ns_r = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

        sheet_rid = None
        for sheet_elem in wb_root.iter(f'{{{ns_main}}}sheet'):
            if sheet_elem.get('name') == sheet_name:
                sheet_rid = sheet_elem.get(f'{{{ns_r}}}id')
                break

        if not sheet_rid:
            return None

        # workbook.xml.rels からrId→sheetN.xmlの対応を取得
        wb_rels = z.read('xl/_rels/workbook.xml.rels')
        rels_root = ET.fromstring(wb_rels)
        ns_rel = 'http://schemas.openxmlformats.org/package/2006/relationships'

        sheet_file = None
        for rel in rels_root.iter(f'{{{ns_rel}}}Relationship'):
            if rel.get('Id') == sheet_rid:
                sheet_file = rel.get('Target')
                break

        if not sheet_file:
            return None

        # sheetN.xml のrels からdrawingファイルを取得
        sheet_basename = sheet_file.split('/')[-1]
        rels_path = f'xl/worksheets/_rels/{sheet_basename}.rels'

        if rels_path not in z.namelist():
            return None

        sheet_rels = z.read(rels_path)
        sheet_rels_root = ET.fromstring(sheet_rels)

        drawing_files = []
        for rel in sheet_rels_root.iter(f'{{{ns_rel}}}Relationship'):
            target = rel.get('Target', '')
            if 'drawing' in target and target.endswith('.xml'):
                # 相対パス→絶対パスに変換
                if target.startswith('../'):
                    target = 'xl/' + target[3:]
                elif not target.startswith('xl/'):
                    target = 'xl/drawings/' + target
                drawing_files.append(target)

        return drawing_files if drawing_files else None

    except Exception as e:
        logger.debug(f'シート-Drawing対応取得エラー: {e}')
        return None


# ──────────────────────────────────────────
# Step 8: 稼働チェック（土日祝・15分単位・深夜残業）
# ──────────────────────────────────────────

def check_work_alerts(daily_data, holiday_dates=None):
    """
    稼働データの包括チェックを行う。

    チェック項目:
      1. 土日・祝日の稼働
      2. 稼働時間の15分単位チェック
      3. 開始・終了時刻の15分単位チェック
      4. 深夜残業チェック（22時以降の終了）

    Args:
        daily_data: 日別データリスト
        holiday_dates: Excel内の祝日シートから取得した{date: name}の辞書

    Returns:
        list: 警告リスト
    """
    if holiday_dates is None:
        holiday_dates = {}

    alerts = []
    for entry in daily_data:
        hours = entry['hours']
        if hours <= 0:
            continue

        d = date.fromisoformat(entry['date'])
        start_str = entry.get('start', '')
        end_str = entry.get('end', '')

        # ── 1. 土日チェック ──
        if d.weekday() >= 5:
            day_type = '土曜' if d.weekday() == 5 else '日曜'
            alerts.append({
                'date': entry['date'],
                'type': 'weekend',
                'hours': hours,
                'day_name': entry.get('day_name', day_type),
                'holiday_name': day_type,
            })

        # ── 2. 祝日チェック ──
        holiday_name = holiday_dates.get(d)
        if holiday_name is None:
            if jpholiday.is_holiday(d):
                holiday_name = jpholiday.is_holiday_name(d)
        if holiday_name:
            alerts.append({
                'date': entry['date'],
                'type': 'holiday',
                'hours': hours,
                'day_name': entry.get('day_name', ''),
                'holiday_name': holiday_name,
            })

        # ── 3. 稼働時間の15分単位チェック ──
        # 15分 = 0.25時間 の倍数でない場合に警告
        remainder = round(hours % 0.25, 4)
        if remainder != 0:
            alerts.append({
                'date': entry['date'],
                'type': 'time_unit',
                'hours': hours,
                'day_name': entry.get('day_name', ''),
                'detail': f'稼働時間 {hours}h が15分単位ではありません',
            })

        # ── 4. 開始・終了時刻の15分単位チェック ──
        for label, time_str in [('開始', start_str), ('終了', end_str)]:
            if not time_str:
                continue
            try:
                parts = time_str.split(':')
                minute = int(parts[1]) if len(parts) > 1 else 0
                if minute % 15 != 0:
                    alerts.append({
                        'date': entry['date'],
                        'type': 'time_unit',
                        'hours': hours,
                        'day_name': entry.get('day_name', ''),
                        'detail': f'{label}時刻 {time_str} が15分単位ではありません',
                    })
            except (ValueError, IndexError):
                pass

        # ── 5. 深夜残業チェック（22時以降の終了）──
        if end_str:
            try:
                end_hour = int(end_str.split(':')[0])
                if end_hour >= 22:
                    alerts.append({
                        'date': entry['date'],
                        'type': 'late_night',
                        'hours': hours,
                        'day_name': entry.get('day_name', ''),
                        'detail': f'終了 {end_str} → 深夜残業の可能性',
                    })
            except (ValueError, IndexError):
                pass

    return alerts


# 後方互換
def check_weekend_holiday_work(daily_data, holiday_dates=None):
    """後方互換: check_work_alertsのエイリアス"""
    return check_work_alerts(daily_data, holiday_dates)


def _parse_holiday_sheet(wb):
    """Excel内の祝日シートから祝日データを読み取る。"""
    holiday_dates = {}

    for sheet_name in wb.sheetnames:
        if '祝日' in sheet_name or 'holiday' in sheet_name.lower():
            ws = wb[sheet_name]
            for row_idx in range(1, min(ws.max_row + 1, 100)):
                for col_idx in range(1, 5):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if isinstance(cell.value, datetime):
                        d = cell.value.date()
                        # 隣のセルから祝日名を取得
                        name_cell = ws.cell(row=row_idx, column=col_idx + 1)
                        name = str(name_cell.value or '祝日')
                        holiday_dates[d] = name
            break

    return holiday_dates
